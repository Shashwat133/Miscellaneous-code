#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import openpyxl
import re 
from googletrans import Translator
import csv
trans = Translator()


# In[2]:


df=pd.read_csv('D:\job\ministryoffinance.csv')


# In[3]:


df1=('D:\job\subtopics.xlsx')


# In[4]:


transl = []
count = 0
for element in df['headline']:
    count += 1
    print("element-" + str(count) + ": " + str(element))
    if(element != ''):
        count += 1
        # check if String is in English or Not
        try:
            readitem = str(element)
            readitem.encode('ascii')
        except UnicodeEncodeError:
            #print("-->>>>Hedline is not in English...")
            # REINITIALIZE THE API
            trans = Translator()
            translatedtext = trans.translate(str(element)).text
            transl.append(translatedtext)
            print(translatedtext)
        else:
            #print("<<--Hedline is in English...")
            transl.append(element)
        ### end Checking...###


# In[5]:


df['trans']=transl


# In[9]:


#reading the list from one file
file=[]
obj = openpyxl.load_workbook(df1) 
sobj = obj.active 
mrow = sobj.max_row 
for i in range(1, mrow + 1): 
    cobj = sobj.cell(row = i, column = 1) 
    print(cobj.value)
    #print the topics
    a=[cobj.value]
    for element in df['trans']:
        if (element !=''):
            #check for the topic
            try:
                readitem=str(element)
                read = str(cobj.value)
            except UnicodeEncodeError:
                #res = [x for x in readitem if re.search(read, x)] 
                res = list(filter(lambda x: read in x, readitem))
                file.append(res)
                print (str(res))
            else:
                print("thank you")
                file.append(element)


# In[50]:


GST=[]
GST=df[df['trans'].str.contains(str('GST'))]


# In[10]:


pd.DataFrame(GST).to_csv("D:\job\demo10.csv")

